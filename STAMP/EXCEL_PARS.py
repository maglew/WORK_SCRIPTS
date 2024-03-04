import codecs

import pandas as pd
from openpyxl import load_workbook

f = codecs.open('input', 'r', 'utf-8')
inp = f.read()
lines = inp.split('\n')

wb = load_workbook('./extract_base_params_construction_v01.xlsx')

sheet = wb['Лист1']

sPARENT = 'EXTRACT_BASE_PARAMS_UNDER_CONSTRUCTION'
output = open('output.txt', 'w')
n = sheet.max_row
tag = " "
text = " "
rownum = " "
type = " "
format = " "
group = " "
note = ""

arr0 = [['','','','','','']]

arr1 = [['','','','','','',]]


sPARENT = sPARENT.upper()



for i in range(1,n+1):
    rownum = str(i)
    group = sheet['A' + rownum].value
    tag = sheet['B' + rownum].value
    type = sheet['C' + rownum].value
    format = sheet['D' + rownum].value
    text = sheet['E' + rownum].value
    note = sheet['F' + rownum].value
    arr0.append([group,tag,type,format,text,note])


master = ''


def findtype(type):
    outtype = ''
    for k in range(0,len(arr0)):
        if (arr0[k][5] is not None and arr0[k][5].find(' '+type+'.') != -1):
            outtype = arr0[k][1]
            return outtype
    return None


for j in range(0,len(arr0)):
    #rownum = str(j)
    tempmaster = arr0[j][0]
    tag = arr0[j][1]
    type = arr0[j][2]
    format = arr0[j][3]
    text = arr0[j][4]
    note = arr0[j][5]
    if (tempmaster is not None):
        if(tempmaster[0:3] == 'Тип'):
            master = findtype(tempmaster[4:tempmaster.find('(')-1])
            if (master is None):
                master = findtype(tempmaster[4:])
            if (master is not None):
                master = master.replace(' ','')
        elif(tempmaster.find('_') != -1):
            master = tempmaster[:tempmaster.find('(')]

    #print(arr0[j])
    if tag is not None and type is not None and tag != 'Содержание элемента':
        arr1.append([master, tag, type, format, text, note,0])
        #output.write(master+']['+ tag+']['+ type+']['+ format+']['+  text+']['+  note + "\n")  #

        #print(arr1[-1])


# вывод отстутствия тегов
'''str = ''
for i in lines:
    str = i.replace('\r','')
'''
#print(inp.find('HEADER'))
index = 0
usl = ''
for i in range(0,len(arr1)):
    master = arr1[i][0]
    tag = arr1[i][1]
    type = arr1[i][2]
    format = arr1[i][3]
    text = arr1[i][4]
    note = arr1[i][5]

    if (master is not None):
        sPARENT = master.upper()
        sPARENT = sPARENT.replace(' ','')

    if (tag is not None):
        tag = tag.upper()

    if (tag is not None and sPARENT is not None):
        usl = sPARENT+'.'+tag
        index = inp.find(usl)
        if  (index  == -1):
            output.write("  отсутствует узел[" + usl +"] ("+ format +") -- "+text+"\n")
        #else:
        #    output.write("  найден узел[" + usl + "] "+text+"\n")

'''
#вывод несорт разпознаний
for i in range(0,len(arr1)):
    master = arr1[i][0]
    tag = arr1[i][1]
    type = arr1[i][2]
    format = arr1[i][3]
    text = arr1[i][4]
    note = arr1[i][5]

    if (master is not None):
        sPARENT = master.upper()
        sPARENT = sPARENT.replace(' ','')

    if (tag is not None):
        tag = tag.upper()

    if (tag is not None and sPARENT is not None):
        output.write("when '" + sPARENT + "." + tag + "' then -- " + "[" + type + "/" + format + "] " + text + "\n")  #

        if (format == 'S'):
            output.write("  BDOWN := true;" + "\n")
        else:
            output.write("  BDOWN := false;" + "\n")
            output.write("  RECORD_" + tag + ";" + "\n")
        output.write(" " + "\n")  #

output.write("----------------------------" + "\n")  #




#вывод функций
for i in range(0,len(arr1)):
    master = arr1[i][0]
    tag = arr1[i][1]
    type = arr1[i][2]
    format = arr1[i][3]
    text = arr1[i][4]
    note = arr1[i][5]

    if (master is not None):
        sPARENT = master.upper()
        sPARENT = sPARENT.replace(' ','')

    if (tag is not None):
        tag = tag.upper()
    if (format == 'Т' or format == 'К'):
        output.write("procedure  RECORD_" + tag + " as" + "\n")
        output.write("begin" + "\n")
        output.write("s" + tag + " := PKG_XPATH.VALUE(L_NODE);" +  "\n")
        output.write("end; --" + text + "\n")
        output.write(" " + "\n")  #
    elif(format == 'N' or format == 'B' or format == 'Z'):
        output.write("procedure  RECORD_" + tag + " as" + "\n")
        output.write("begin" + "\n")
        output.write("n" + tag + " := F_XML_NUMBER_FROM_STR(PKG_XPATH.VALUE(L_NODE));" + "\n")
        output.write("end; --" + text + "\n")
        output.write(" " + "\n")  #
    elif(format == 'D' or format == 'DT'):
        output.write("procedure  RECORD_" + tag + " as" + "\n")
        output.write("begin" + "\n")
        output.write("d" + tag + " := STR2DAT(PKG_XPATH.VALUE(L_NODE));" + "\n")
        output.write("end; --" + text + "\n")
        output.write(" " + "\n")  #

#вывод переменных
for i in range(0,len(arr1)):
    master = arr1[i][0]
    tag = arr1[i][1]
    type = arr1[i][2]
    format = arr1[i][3]
    text = arr1[i][4]
    note = arr1[i][5]

    if (master is not None):
        sPARENT = master.upper()
        sPARENT = sPARENT.replace(' ','')

    if (tag is not None):
        tag = tag.upper()
    if (format == 'Т' or format == 'К'):
        #SR_MUNICIP     PKG_STD.TSTRING;
        output.write("S" + tag + "      PKG_STD.TSTRING; --" + text + "\n")
    elif(format == 'N' or format == 'B' or format == 'Z'):
        #NTYPEDOC      PKG_STD.TNUMBER;
        output.write("N" + tag + "      PKG_STD.TNUMBER; --" + text +"\n")

    elif(format == 'D' or format == 'DT'):
        #D_sdsd date;
        output.write("D" + tag + "      date; --" + text + "\n")
'''
'''
#вывод сортированных распознаний
for i in range(0,len(arr1)):
    master = arr1[i][0]
    tag = arr1[i][1]
    type = arr1[i][2]
    format = arr1[i][3]
    text = arr1[i][4]
    note = arr1[i][5]
    lvl = arr1[i][6]

    if (master is not None):
        sPARENT = master.upper()
        sPARENT = sPARENT.replace(' ','')

    if (tag is not None):
        tag = tag.upper()
    if (lvl == 0 and format == 'S'):
        print(print(arr1[i]))
    if (lvl == 0 and format != 'S'):
        print(print(arr1[i]))



for i in range(1,n+1):
    rownum =  str(i)
    group = sheet['A'+rownum].value
    tag = sheet['B'+rownum].value
    type = sheet['C' + rownum].value
    format = sheet['D' + rownum].value
    text = sheet['E'+rownum].value
    note = sheet['D' + rownum].value
    if (tag is not None):
        tag = tag.upper()

    if (group is not None and group.find('(') != -1 and group.find('Тип') == -1):
        sPARENT = group[:group.find('(')]
        sPARENT = sPARENT.replace(' ','')
        sPARENT = sPARENT.upper()

    if  (tag is not None and sPARENT is not None):
        output.write("when '"+sPARENT+"."+tag+"' then -- " +"["+type+"/"+format+"] "+text+ "\n")  #

        if (format == 'S'):
            output.write("  BDOWN := true;" + "\n")
            output.write("---------------------------" + "\n")
        else:
            output.write("  BDOWN := false;" + "\n")
            output.write("  RECORD_"+tag+";" + "\n")
        output.write(" " + "\n")  #
'''