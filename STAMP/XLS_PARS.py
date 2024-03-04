import codecs

import pandas as pd
from openpyxl import load_workbook

f = codecs.open('input.txt', 'r', 'utf-8')
inp = f.read()
lines = inp.split('\n')

wb = load_workbook('./doc.xlsx')

sheet = wb['Лист1']

output = open('output.txt', 'w')
n = sheet.max_row
SUNITNAME = []
SUNITNAME_TMP = []
SUNITREC = []
arr0 = [['','','','','']]
name = ""
code = ""
table = ""
control1 =""
control2 =""
'''
tablenames = \
    [['Исполнения должностей','CLNPSPFM'],
     ['Приказы (пункты. параметры исполнения)', 'PRORDSPPFM'],
     ['Исполнения должностей(хроника)','CLNPSPFMHS'],
     ['Входящие ведомости (сотрудники)','SLINCSHTSP'],
     ['Сотрудники','CLNPERSONS'],
     ['Сотрудники (исполнительные листы)','CLNPERSEXECACT'],
     ['Журнал больничных листов (уход за родственниками)','PRMEDCERREL'],
     ['Сведения для выплаты пособий','FSSPAYM'],
     ['Сотрудники (положенные вычеты)','CLNPERSDED'],
     ['Журнал отклонений','TBDEVIAT'],
     ['Штатные должности','CLNPSDEP'],
     ['Приказы (пункты. параметры штатной должности)','PRORDSPPOST'],
     ['Приказы (пункты. параметры подразделения)','sss']]


for i in range(1,n+1):
    rownum = str(i)
    name = sheet['A' + rownum].value
    code = sheet['B' + rownum].value
    table = sheet['C' + rownum].value
    control1 = sheet['D' + rownum].value
    control2 = sheet['E' + rownum].value
    arr0.append([name,code,table,control1,control2])


for line in arr0:
    if (line[4] != '' and line[4] != '-' and line[4] != 5):
        #print(line)
        SUNITNAME_TMP = line[4].split(',')
        #print(SUNITNAME_TMP)
        for tmp in SUNITNAME_TMP:
            if (SUNITNAME.count(tmp.strip()) == 0):
                SUNITNAME.append(tmp.strip())
                #SUNITREC.append([tmp.strip(),line[2]])
print(SUNITNAME)

stable = ''

for unit in SUNITNAME:
    output.write("/* Проверка актуальности связей у раздела <<" + unit + ">>*/\n")

    #stable = tablenames[:1]
    for table in tablenames:
        if (table[0]==unit):
            stable = table[1]
    output.write("elsif STABLE = '" + stable + "' then \n")
    for line in arr0:
        if (line[4] != '' and line[4] != '-' and line[4] != 5 and line[4].find(unit)!=-1):
            #output.write("  /* RN с таблицы  "  + line[0] + " */ \n")
            output.write("  STEMPUNICODE := "+line[1] +";\n")
            output.write("  begin\n")
            output.write("      select T./*"+line[2] +"*/ into NREC_RN from " + stable + " T where T.RN = NRN; --  Проверка cвязей с разделом <<"+line[0]+">>  \n")
            output.write("  exception when NO_DATA_FOUND then\n")
            output.write("      NREC_RN := null;\n")
            output.write("  end;\n")
            output.write("  if NREC_RN is not null then\n")
            output.write("      UDO_P_RECORD_DELITION_CHECK\n")
            output.write("      (\n")
            output.write("          NTYPE_ERR   => 2, \n")
            output.write("          NREC_RN     => NREC_RN,\n")
            output.write("          SUNITCODE   => STEMPUNICODE\n")
            output.write("      );\n")
            output.write("  end if;\n")
            output.write("\n")


    #output.write("\n")
output.write("END if;\n")
'''

STABLE = ['AGNADDRESSES', 'CLNPSPFMWD', 'CLNPSPFMWH', 'AGNDOCUMS', 'CLNPERSDED', 'AGNRELATIVE', 'AGNDISABLED', 'CLNPERSEXP', 'CLNPSPFMGS']

for tab in STABLE:
    output.write("elsif STABLE = '" + tab + "' then \n")
    output.write("  begin\n")
    output.write("      select T.PRN into NPRN from " + tab + " T where T.RN = NRN; \n")
    output.write("  exception when NO_DATA_FOUND then\n")
    output.write("      p_exception(0,'Родитель у раздела с таблицей "+tab+" и РН '||NRN||' не найден');\n")
    output.write("  end;\n")
output.write("END if;\n")